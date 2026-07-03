from __future__ import annotations

from io import BytesIO

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill


def read_uploaded_file(uploaded_file) -> pd.DataFrame:
    """Read CSV or XLSX file from Streamlit uploader into a DataFrame."""
    file_name = (uploaded_file.name or "").lower()

    if file_name.endswith(".csv"):
        return pd.read_csv(uploaded_file)

    if file_name.endswith(".xlsx"):
        return pd.read_excel(uploaded_file, engine="openpyxl")

    raise ValueError("Unsupported file type. Please upload a CSV or XLSX file.")


def validate_url_column(df: pd.DataFrame) -> None:
    """Ensure required URL column exists in uploaded input."""
    if "URL" not in df.columns:
        raise ValueError('Input file must contain a column named "URL".')


def results_to_excel_bytes(df: pd.DataFrame) -> bytes:
    """Serialize DataFrame as XLSX bytes and highlight sentiment mismatches in yellow."""
    buffer = BytesIO()
    df.to_excel(buffer, index=False, engine="openpyxl")

    buffer.seek(0)
    workbook = load_workbook(buffer)
    worksheet = workbook.active

    header_map = {cell.value: cell.column for cell in worksheet[1]}
    comparison_col = header_map.get("Sentiment Comparison")
    if comparison_col is not None:
        yellow_fill = PatternFill(start_color="FFF59D", end_color="FFF59D", fill_type="solid")
        for row_idx in range(2, worksheet.max_row + 1):
            cell = worksheet.cell(row=row_idx, column=comparison_col)
            if str(cell.value).strip().lower() == "mismatch":
                cell.fill = yellow_fill

    output = BytesIO()
    workbook.save(output)
    output.seek(0)
    return output.getvalue()


def get_input_sentiments(df: pd.DataFrame) -> list[str]:
    """Return input sentiments row-wise; defaults to blank when column is absent."""
    if "Sentiment" not in df.columns:
        return [""] * len(df)

    return df["Sentiment"].fillna("").astype(str).tolist()


def has_sentiment_column(df: pd.DataFrame) -> bool:
    """Check whether uploaded input contains the Sentiment column."""
    return "Sentiment" in df.columns
